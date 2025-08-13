import os
import shutil
import pandas as pd
import numpy as np
import logging
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import calendar
from pathlib import Path
from typing import Optional, List
import sys

# -------------------- CONFIGURACIÓN DE LOGGING --------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('proyeccion_caja.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# -------------------- CONFIGURACIÓN --------------------
class Config:
    """Clase para manejar la configuración del sistema"""
    
    def __init__(self):
        self.fecha_actual = datetime.now().strftime("%d%m%y")
        self.hoy = date.today()
        self.inicio_mes = self.hoy.replace(day=1)
        ultimo_dia = calendar.monthrange(self.hoy.year, self.hoy.month)[1]
        self.fin_mes = self.hoy.replace(day=ultimo_dia)
        self.fecha_inicio = pd.to_datetime(self.inicio_mes)
        self.fecha_fin = pd.to_datetime(self.fin_mes)
        
        # Rutas
        self.PATH_FILES = Path(r"C:\Users\ntorreslo\Downloads")
        self.DESTINO_BASE = Path(r"\\192.168.235.35\GerenciaMercadoCapitales\02_Riesgo_Deuda\01_Back_Office\2_Proyección_de_Caja\2025\8. Agosto")
        self.TEMPLATE_EXCEL = Path(r"\\192.168.235.35\GerenciaMercadoCapitales\02_Riesgo_Deuda\01_Back_Office\2_Proyección_de_Caja\Plantilla Proyección de caja.xlsx")
        self.carpeta_destino = self.DESTINO_BASE / self.fecha_actual
        
        # Prefijos de archivos a copiar
        self.PREFIJOS = ["Liquidaciones_NDF", "Intereses_de_Deuda", "Detalle_Swap", "NDF_Vigentes_y_Vtos"]
        
        # Configuración de hojas Excel
        self.CONFIGURACION_HOJAS = {
            "Liquidaciones_NDF": {"hoja_origen": "DATOS", "hoja_destino": "NDF_data"},
            "NDF_Vigentes_y_Vtos": {"hoja_origen": "DATOS2", "hoja_destino": "Clasificacion_NDF"},
            "Intereses_de_Deuda": {"hoja_origen": "DATOS", "hoja_destino": "Deuda_data"},
            "Detalle_Swap": {"hoja_origen": "DATOS", "hoja_destino": "SWAPS_data"}
        }

# -------------------- FUNCIONES AUXILIARES --------------------

def validar_rutas(config: Config) -> bool:
    """Valida que las rutas necesarias existan"""
    try:
        if not config.PATH_FILES.exists():
            logger.error(f"La carpeta de descargas no existe: {config.PATH_FILES}")
            return False
            
        if not config.DESTINO_BASE.exists():
            logger.error(f"La carpeta de destino base no existe: {config.DESTINO_BASE}")
            return False
            
        if not config.TEMPLATE_EXCEL.exists():
            logger.error(f"La plantilla de Excel no existe: {config.TEMPLATE_EXCEL}")
            return False
            
        return True
    except Exception as e:
        logger.error(f"Error al validar rutas: {e}")
        return False

def crear_carpeta_destino(carpeta_destino: Path) -> bool:
    """Crea la carpeta de destino si no existe"""
    try:
        if carpeta_destino.exists():
            logger.info(f"La carpeta {carpeta_destino} ya existe. Continuando con el procesamiento...")
            return True
        else:
            carpeta_destino.mkdir(parents=True, exist_ok=True)
            logger.info(f"Carpeta creada: {carpeta_destino}")
            return True
    except Exception as e:
        logger.error(f"Error al crear carpeta de destino: {e}")
        return False

def archivo_mas_reciente(prefijo: str, carpeta: Path) -> Optional[Path]:
    """Encuentra el archivo más reciente con el prefijo dado"""
    try:
        archivos = list(carpeta.glob(f"{prefijo}*.xlsx"))
        if not archivos:
            logger.warning(f"No se encontraron archivos con prefijo '{prefijo}' en {carpeta}")
            return None
        
        archivo_reciente = max(archivos, key=lambda f: f.stat().st_mtime)
        logger.info(f"Archivo más reciente encontrado: {archivo_reciente.name}")
        return archivo_reciente
    except Exception as e:
        logger.error(f"Error al buscar archivo con prefijo '{prefijo}': {e}")
        return None

def copiar_archivos(config: Config) -> List[str]:
    """Copia los archivos necesarios a la carpeta de destino"""
    archivos_copiados = []
    
    # Usar getattr con fallback para evitar el AttributeError
    prefijos = getattr(config, 'PREFIJOS', [
        "Liquidaciones_NDF", 
        "Intereses_de_Deuda", 
        "Detalle_Swap", 
        "NDF_Vigentes_y_Vtos"
    ])
    
    logger.info(f"Prefijos a procesar: {prefijos}")
    
    try:
        for prefijo in prefijos:
            archivo = archivo_mas_reciente(prefijo, config.PATH_FILES)
            if archivo:
                destino = config.carpeta_destino / archivo.name
                shutil.copy2(archivo, destino)
                archivos_copiados.append(archivo.name)
                logger.info(f"Archivo '{archivo.name}' copiado exitosamente")
            else:
                logger.warning(f"No se encontró archivo con prefijo '{prefijo}'")
        
        return archivos_copiados
    except Exception as e:
        logger.error(f"Error al copiar archivos: {e}")
        return archivos_copiados

def copiar_y_renombrar_plantilla(config: Config) -> Optional[Path]:
    """Copia y renombra la plantilla de Excel"""
    try:
        plantilla_destino = config.carpeta_destino / f"Proyección de caja {config.fecha_actual}.xlsx"
        
        # Si ya existe la plantilla renombrada, no hacer nada
        if plantilla_destino.exists():
            logger.info(f"La plantilla ya existe en destino: {plantilla_destino}")
            return plantilla_destino
        
        # Copiar plantilla
        shutil.copy2(config.TEMPLATE_EXCEL, config.carpeta_destino)
        logger.info("Plantilla de Excel copiada")
        
        # Renombrar
        plantilla_original = config.carpeta_destino / "Plantilla Proyección de caja.xlsx"
        if plantilla_original.exists():
            plantilla_original.rename(plantilla_destino)
            logger.info("Plantilla renombrada correctamente")
            return plantilla_destino
        else:
            logger.error("No se encontró la plantilla para renombrar")
            return None
            
    except Exception as e:
        logger.error(f"Error al copiar y renombrar plantilla: {e}")
        return None

def limpiar_hoja_excel(ws) -> None:
    """Limpia el contenido de una hoja de Excel manteniendo los encabezados"""
    try:
        max_row = ws.max_row if ws.max_row else 1
        max_col = ws.max_column if ws.max_column else 1
        
        for row in ws.iter_rows(min_row=2, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = None
        logger.debug(f"Hoja '{ws.title}' limpiada correctamente")
    except Exception as e:
        logger.error(f"Error al limpiar hoja '{ws.title}': {e}")

def escribir_dataframe_en_hoja(ws, df: pd.DataFrame) -> None:
    """Escribe un DataFrame en una hoja de Excel"""
    try:
        if df.empty:
            logger.warning(f"DataFrame vacío para la hoja '{ws.title}'")
            return
            
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                # Manejar valores None o NaN
                if pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        logger.info(f"DataFrame escrito en hoja '{ws.title}': {len(df)} filas")
    except Exception as e:
        logger.error(f"Error al escribir DataFrame en hoja '{ws.title}': {e}")

def procesar_liquidaciones_ndf(config: Config, wb) -> None:
    """Procesa los datos de Liquidaciones NDF"""
    try:
        ruta_ndf = archivo_mas_reciente("Liquidaciones_NDF", config.PATH_FILES)
        if not ruta_ndf:
            logger.warning("No se encontró archivo de Liquidaciones_NDF")
            return
        
        df_ndf = pd.read_excel(ruta_ndf, sheet_name="DATOS", engine="openpyxl")
        df_ndf['F_VENCI'] = pd.to_datetime(df_ndf['F_VENCI'], errors='coerce')
        df_ndf = df_ndf.sort_values(by='F_VENCI', ascending=True)
        
        
        # Hacer fechas únicas para los NDF
        df_ndf['F_VENCI2'] = df_ndf['F_VENCI'] +timedelta(days=1)
        fechas_ndf = df_ndf['F_VENCI2'].to_frame(name='Fechas').drop_duplicates().sort_values(by = 'Fechas')

        # Guardar datos para clasificación posterior
        global df_codigo
        df_codigo = df_ndf[['CODIGO', 'OBSERVACIÓN']].rename(
            columns={'CODIGO': 'BOLETA', 'OBSERVACIÓN': 'OBSERVACION'}
        )
        
        ws = wb["NDF_data"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, df_ndf)

        # Escribir las fechas en la hoja correspondiente
        ws = wb["Fecha_NDF"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, fechas_ndf)


        logger.info("Liquidaciones NDF procesadas correctamente")
        
    except Exception as e:
        logger.error(f"Error al procesar Liquidaciones NDF: {e}")

def clasificar_tipo(row) -> Optional[str]:
    """Clasifica el tipo de NDF basado en los códigos y observaciones"""
    try:
        cod_instr = row.get('COD INSTRUMENTO', '')
        opexcapex = row.get('OPEXCAPEX', '')
        observacion = str(row.get('OBSERVACION', '')).lower()
        
        if cod_instr == 'NDC':
            return 'NDF Capex'
        elif cod_instr == 'NDF':
            return 'NDF Financiero'
        elif cod_instr == 'NDO':
            return 'NDF Opex'
        elif cod_instr == 'NDP':
            if opexcapex == 'OPEX':
                return 'NDF Opex'
            elif opexcapex == 'CAPEX':
                return 'NDF Capex'
        
        if 'capex' in observacion:
            return 'NDF Capex'
        elif 'opex' in observacion:
            return 'NDF Opex'
            
        return None
    except Exception as e:
        logger.error(f"Error al clasificar tipo: {e}")
        return None

def procesar_ndf_vigentes(config: Config, wb) -> None:
    """Procesa los datos de NDF Vigentes y Vencimientos"""
    try:
        global df_codigo
        if 'df_codigo' not in globals():
            logger.warning("No se encontraron datos de código de NDF previos")
            return
            
        ruta_ndf = archivo_mas_reciente("NDF_Vigentes_y_Vtos", config.PATH_FILES)
        if not ruta_ndf:
            logger.warning("No se encontró archivo de NDF_Vigentes_y_Vtos")
            return
        
        df_ndf = pd.read_excel(ruta_ndf, sheet_name="DATOS2", engine="openpyxl")
        df_boletas = df_ndf[['BOLETA', 'COD INSTRUMENTO', 'OPEXCAPEX']]
        df_clasificacion = pd.merge(df_codigo, df_boletas, on='BOLETA', how='left')
        
        df_clasificacion['TIPO'] = df_clasificacion.apply(clasificar_tipo, axis=1)
        df_clasificacion = df_clasificacion[['BOLETA', 'TIPO', 'COD INSTRUMENTO', 'OPEXCAPEX']]
        
        ws = wb["Clasificacion_NDF"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, df_clasificacion)
        logger.info("NDF Vigentes y Vencimientos procesados correctamente")
        
    except Exception as e:
        logger.error(f"Error al procesar NDF Vigentes: {e}")

def procesar_intereses_deuda(config: Config, wb) -> None:
    """Procesa los datos de Intereses de Deuda"""
    try:
        ruta_ndf = archivo_mas_reciente("Intereses_de_Deuda", config.PATH_FILES)
        if not ruta_ndf:
            logger.warning("No se encontró archivo de Intereses_de_Deuda")
            return
        
        df_ndf = pd.read_excel(ruta_ndf, sheet_name="DATOS", engine="openpyxl")
        
        ws = wb["Deuda_data"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, df_ndf)
        logger.info("Intereses de Deuda procesados correctamente")
        
    except Exception as e:
        logger.error(f"Error al procesar Intereses de Deuda: {e}")

def procesar_detalle_swap(config: Config, wb) -> None:
    """Procesa los datos de Detalle Swap"""
    try:
        ruta_ndf = archivo_mas_reciente("Detalle_Swap", config.PATH_FILES)
        if not ruta_ndf:
            logger.warning("No se encontró archivo de Detalle_Swap")
            return
        
        df_ndf = pd.read_excel(ruta_ndf, sheet_name="DATOS", engine="openpyxl")
        df_ndf['F_LIQUI'] = pd.to_datetime(df_ndf['F_LIQUI'], errors='coerce')
        
        # Filtrar por fechas del mes actual
        df_ndf = df_ndf[
            (df_ndf['F_LIQUI'] >= config.fecha_inicio) & 
            (df_ndf['F_LIQUI'] <= config.fecha_fin)
        ]
        
        # Excluir instrumentos que empiecen con NDF
        df_ndf = df_ndf[~df_ndf['INSTRUMENTO'].str.startswith('NDF', na=False)]
        
        # Sacar las fechas únicas de los SWAPS
        fechas_swaps = df_ndf['F_LIQUI'].to_frame(name='Fechas').drop_duplicates().sort_values(by='Fechas')

        # Escribir los datos limpios de los SWAPS
        ws = wb["SWAPS_data"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, df_ndf)

        # Escribir las fechas en la hoja correspondiente
        ws = wb["Fecha_Swap"]
        limpiar_hoja_excel(ws)
        escribir_dataframe_en_hoja(ws, fechas_swaps)


        logger.info("Detalle Swap procesado correctamente")
        
    except Exception as e:
        logger.error(f"Error al procesar Detalle Swap: {e}")


def procesar_datos_excel(config: Config, plantilla_destino: Path) -> bool:
    """Procesa todos los datos y actualiza el archivo Excel"""
    try:
        wb = load_workbook(plantilla_destino)
        logger.info("Archivo Excel abierto para procesamiento")
        
        # Procesar cada tipo de datos
        procesar_liquidaciones_ndf(config, wb)
        procesar_ndf_vigentes(config, wb)
        procesar_intereses_deuda(config, wb)
        procesar_detalle_swap(config, wb)
        
        # Guardar el archivo
        wb.save(plantilla_destino)
        wb.close()
        logger.info("Archivo Excel actualizado y guardado correctamente")
        return True
        
    except Exception as e:
        logger.error(f"Error al procesar datos de Excel: {e}")
        return False

# -------------------- FUNCIÓN PRINCIPAL --------------------

def main():
    """Función principal del programa"""
    try:
        logger.info("=== INICIANDO PROCESAMIENTO DE PROYECCIÓN DE CAJA ===")
        
        # Inicializar configuración
        config = Config()
        logger.info(f"Fecha actual: {config.fecha_actual}")
        logger.info(f"Carpeta destino: {config.carpeta_destino}")
        
        # Validar rutas
        if not validar_rutas(config):
            logger.error("Error en validación de rutas. Abortando proceso.")
            return False
        
        # Crear carpeta de destino
        if not crear_carpeta_destino(config.carpeta_destino):
            logger.error("Error al crear carpeta de destino. Abortando proceso.")
            return False
        
        # Copiar archivos
        archivos_copiados = copiar_archivos(config)
        logger.info(f"Archivos copiados: {len(archivos_copiados)}")
        
        # Copiar y renombrar plantilla
        plantilla_destino = copiar_y_renombrar_plantilla(config)
        if not plantilla_destino:
            logger.error("Error al preparar plantilla de Excel. Abortando proceso.")
            return False
        
        # Procesar datos en Excel
        if procesar_datos_excel(config, plantilla_destino):
            logger.info("=== PROCESAMIENTO COMPLETADO EXITOSAMENTE ===")
            return True
        else:
            logger.error("=== ERROR EN EL PROCESAMIENTO ===")
            return False
            
    except Exception as e:
        logger.error(f"Error crítico en función principal: {e}")
        return False

if __name__ == "__main__":
    success = main()
    if not success:
        sys.exit(1)

