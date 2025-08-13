import requests
import pandas as pd
from io import StringIO
from bs4 import BeautifulSoup
import os
from datetime import datetime, date
from openpyxl import load_workbook

# -------------------- CONFIGURACIÓN --------------------

# Fecha actual y formato
fecha_actual = datetime.now().strftime("%d%m%y")
DESTINO_BASE = r"\\192.168.235.35\GerenciaMercadoCapitales\02_Riesgo_Deuda\01_Back_Office\2_Proyección_de_Caja\2025\8. Agosto"
TEMPLATE_EXCEL = r"\\192.168.235.35\GerenciaMercadoCapitales\02_Riesgo_Deuda\01_Back_Office\2_Proyección_de_Caja\Plantilla Proyección de caja.xlsx"
carpeta_destino = os.path.join(DESTINO_BASE, fecha_actual)
plantilla_destino = os.path.join(carpeta_destino, f"Proyección de caja {fecha_actual}.xlsx")

# === Función para obtener el tipo de cambio EUR/USD desde el BCE ===
def obtener_eurusd():
    url = "https://data-api.ecb.europa.eu/service/data/EXR/D.USD.EUR.SP00.A?format=csvdata"
    try:
        response = requests.get(url, verify=False)
        response.raise_for_status()
        df = pd.read_csv(StringIO(response.text))
        df = df[['TIME_PERIOD', 'OBS_VALUE']]
        df['OBS_VALUE'] = pd.to_numeric(df['OBS_VALUE'], errors='coerce')
        ultima_fila = df.tail(1)
        eurusd = ultima_fila['OBS_VALUE'].values[0]
        print(f"Tasa de cambio EUR/USD: {eurusd}")
        return eurusd, ultima_fila
    except Exception as e:
        print(f"Error al obtener EUR/USD: {e}")
        return None, None

# === Función para obtener la TRM (USD/COP) desde la Superfinanciera ===
def obtener_usdcop():
    url = "https://www.superfinanciera.gov.co/CargaDriver/index.jsp"
    try:
        response = requests.get(url, verify=False)
        response.raise_for_status()
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        tables = soup.find_all('table')

        if not tables:
            print("No se encontraron tablas en la página.")
            return None, None

        df = pd.read_html(StringIO(str(tables[0])))[0]
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
        usdcop = df['Valor'].iloc[0]
        print(f"Tasa de cambio USD/COP: {usdcop}")
        return usdcop, df
    except Exception as e:
        print(f"Error al obtener USD/COP: {e}")
        return None, None

# === Ejecución principal ===

eurusd, df_eurusd = obtener_eurusd()
usdcop, df_usdcop = obtener_usdcop()

    # Mostrar los DataFrames si se desea
if df_eurusd is not None:
    print("\nÚltimo dato EUR/USD:")
    print(eurusd)

if df_usdcop is not None:
    print("\nÚltimo dato USD/COP:")
    print(usdcop)

# Abrir el libro una sola vez
wb = load_workbook(plantilla_destino)
ws = wb["Caja Coberturas"]

# Escribir el valor en la celda B7
ws["B7"] = usdcop
# Escribir el valor en la celda B8
ws["C7"] = eurusd

wb.save(plantilla_destino)
# Seleccionar la hoja de SWAPS
#if 'Caja Coberturas' not in wb.sheetnames:
#    wb.create_sheet('Caja Coberturas')



